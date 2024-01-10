#!/usr/bin/env python
# -*- coding: utf-8 -*-

import gzip
import shutil
import re
import os
import shutil
import pickle
import sys
import subprocess
import paramiko
import openpyxl
import statistics
from datetime import datetime


def simple_check(path, u, name):
    file_lis = []
    gzip_lis = []
    for pathname, dirnames, filenames in os.walk(path):
        for filename in filenames:
            #print(f"filename:{filename}")
            # フィルタ処理
            if filename[:u] == name:
                if filename[-2:] == "gz":
                    gzip_lis.append(filename)
                    continue
                file_lis.append(filename)
    return file_lis, gzip_lis


def extract_gzip(gzip_file_path, extract_file_path, client):
    """
    ZIPファイルを指定されたディレクトリに解凍する関数
    :param zip_file_path: 解凍するZIPファイルのパス
    :param extract_to: 解凍先のディレクトリ
    :param client: SSH（HTTP）クライアント
    """
    try:
        # gzipファイルをリモートからローカルに転送
        with client.open_sftp() as sftp:
            sftp.get(gzip_file_path, extract_file_path + '.gz')

        # ローカルでgzipファイルを解凍
        with gzip.open(extract_file_path + '.gz', 'rb') as f_in, open(extract_file_path, 'wb') as f_out:
            shutil.copyfileobj(f_in, f_out)

    except FileNotFoundError:
        print(f"圧縮ログファイル '{gzip_file_path}' が見つかりません。")


def successful_ssh_connections(log_file_path):
    """
    ログファイルを読み込んでsshのアクセスログのみ抽出する関数
    """
    try:
        with open(log_file_path, "r") as log_file:
            log_data = log_file.read()
            pattern = r".*sshd\[\d+\]:.*"
            #pattern = r".*sshd\[\d+\]: Accepted password for .* from"
            ssh_successful_connections = re.findall(pattern, log_data)
            return ssh_successful_connections
    except FileNotFoundError:
        print(f"ログファイル '{log_file_path}' が見つかりません。")


def successful_http_connections(log_file_path):
    """
    ログファイルを読み込んでsshのアクセスログのみ抽出する関数
    """
    try:
        with open(log_file_path, "r") as log_file:
            log_data = log_file.read()
            pattern = r".*GET.*?HTTP.\d+.\d+"
            http_successful_connections = re.findall(pattern, log_data)
            return http_successful_connections
    except FileNotFoundError:
        print(f"ログファイル '{log_file_path}' が見つかりません。")


def ssh_log_extraction(ssh_successful_lis):
    """
    sshログリストから日付のみのリストにして返す関数
    """
    log_date_lis = []
    for i in ssh_successful_lis:
        date = i[0:6]
        log_date_lis.append(date)
    return log_date_lis


def http_log_extraction(http_successful_lis):
    """
    httpログリストから日付のみのリストにして返す関数
    """
    log_date_lis = []
    for i in http_successful_lis:
        date = re.findall(r"[0-9]{1,2}/[a-xA-Z]{1,4}/[0-9]{1,4}", i)
        log_date_lis.append(date[0])
    return log_date_lis


def ssh_sort_date_lis(file_name_lis, dir_path, date_lis):
    all_date_lis = date_lis.copy()
    num = 0
    for i in file_name_lis:
        num += 1
        log_file_path = dir_path + i
        ssh_successful_lis = successful_ssh_connections(log_file_path)
        log_date_lis = ssh_log_extraction(ssh_successful_lis)
        if log_date_lis == []:
            continue
        if num == len(file_name_lis):
            most_new_log = log_date_lis.pop(-1)
        dates = sorted(set(log_date_lis), key=log_date_lis.index)
        for s in dates:
            tf = s in all_date_lis
            if tf == False:
                all_date_lis.append(s)
    lis_mon = [[] for i in range(12)]
    for i in all_date_lis:
        match i[0:3]:
            case "Jan":
                lis_mon[8].append(i)
            case "Feb":
                lis_mon[9].append(i)
            case "Mar":
                lis_mon[10].append(i)
            case "Apr":
                lis_mon[11].append(i)
            case "May":
                lis_mon[0].append(i)
            case "Jun":
                lis_mon[1].append(i)
            case "Jul":
                lis_mon[2].append(i)
            case "Aug":
                lis_mon[3].append(i)
            case "Sep":
                lis_mon[4].append(i)
            case "Oct":
                lis_mon[5].append(i)
            case "Nov":
                lis_mon[6].append(i)
            case "Dec":
                lis_mon[7].append(i)
    for i in range(len(lis_mon)):
        lis_mon[i] = sorted(lis_mon[i])
    sort_date_lis = []
    for i in range(len(lis_mon)):
        sort_date_lis.extend(lis_mon[i])
    return sort_date_lis


def http_sort_date_lis(file_name_lis, dir_path, date_lis):
    all_date_lis = date_lis.copy()
    for i in file_name_lis:
        log_file_path = dir_path + i
        http_successful_lis = successful_http_connections(log_file_path)
        log_date_lis = http_log_extraction(http_successful_lis)
        if log_date_lis == []:
            continue
        dates = sorted(set(log_date_lis), key=log_date_lis.index)
        for s in dates:
            tf = s in all_date_lis
            if tf == False:
                all_date_lis.append(s)
    lis_mon = [[] for i in range(12)]
    for i in all_date_lis:
        match i[3:6]:
            case "Jan":
                lis_mon[8].append(i)
            case "Feb":
                lis_mon[9].append(i)
            case "Mar":
                lis_mon[10].append(i)
            case "Apr":
                lis_mon[11].append(i)
            case "May":
                lis_mon[0].append(i)
            case "Jun":
                lis_mon[1].append(i)
            case "Jul":
                lis_mon[2].append(i)
            case "Aug":
                lis_mon[3].append(i)
            case "Sep":
                lis_mon[4].append(i)
            case "Oct":
                lis_mon[5].append(i)
            case "Nov":
                lis_mon[6].append(i)
            case "Dec":
                lis_mon[7].append(i)
    for i in range(len(lis_mon)):
        lis_mon[i] = sorted(lis_mon[i])
    sort_date_lis = []
    for i in range(len(lis_mon)):
        sort_date_lis.extend(lis_mon[i])
    return sort_date_lis


def directory_rm(directory_path):
    try:
        shutil.rmtree(directory_path)
        print(f"The directory '{directory_path}' has been successfully removed.")
    except OSError as e:
        print(f"directory_rm--Error: {e}")


def ssh_Average_access_frequency_calculation(log_date_lis, ssh_successful_count):
    """
    平均アクセス頻度を返す関数
    """
    total_difference_date = 0
    total_difference_date_lis = []
    try:
        for i in range(len(log_date_lis)-1):
            if log_date_lis[i][0:2] == log_date_lis[i+1][0:2]:
                difference_date = int(log_date_lis[i+1][-2:]) - int(log_date_lis[i][-2:])
            else:
                if log_date_lis[0:2] == 'Jan' or 'Mar' or 'May' or 'Jul' or 'Aog' or 'Oct' or 'Dec':
                    difference_date = 31 - int(log_date_lis[i][-2:]) + int(log_date_lis[i+1][-2:])
                else:
                    difference_date = 30 - int(log_date_lis[i][-2:]) + int(log_date_lis[i+1][-2:])
            total_difference_date_lis.append(difference_date)
            total_difference_date += difference_date
        print(total_difference_date_lis, total_difference_date)
        Average_access_frequency = total_difference_date / (ssh_successful_count - 1)
    except ZeroDivisionError:
         Average_access_frequency = 0
    finally:
        return Average_access_frequency

def ssh_maximum_value_access_frequency(log_date_lis):
    total_difference_date_lis = []
    try:
        print(f"ログ：{log_date_lis}")
        for i in range(len(log_date_lis)-1):
            if log_date_lis[i][0:2] == log_date_lis[i+1][0:2]:
                difference_date = int(log_date_lis[i+1][-2:]) - int(log_date_lis[i][-2:])
            else:
                if log_date_lis[0:2] == 'Jan' or 'Mar' or 'May' or 'Jul' or 'Aog' or 'Oct' or 'Dec':
                    difference_date = 31 - int(log_date_lis[i][-2:]) + int(log_date_lis[i+1][-2:])
                else:
                    difference_date = 30 - int(log_date_lis[i][-2:]) + int(log_date_lis[i+1][-2:])
            total_difference_date_lis.append(difference_date)
        print(total_difference_date_lis)
        total_difference_date_lis = sorted(total_difference_date_lis)
        maximum_value = total_difference_date_lis[-1]
    except OSError as e:
        print(f"maximum_value--Error: {e}")
        maximum_value = 100
    finally:
        return  maximum_value

def http_maximum_value_access_frequency(log_date_lis):
    total_difference_date_lis = []
    print(log_date_lis)
    try:
        for i in range(len(log_date_lis)-1):
            if log_date_lis[i][3:6] == log_date_lis[i+1][3:6]:
                difference_date = int(log_date_lis[i+1][:2]) - int(log_date_lis[i][:2])
            else:
                if log_date_lis[3:6] == 'Jan' or log_date_lis[3:6] == 'Mar' or log_date_lis[3:6] == 'May' or log_date_lis[3:6] == 'Jul' or log_date_lis[3:6] == 'Aug' or log_date_lis[3:6] == 'Oct' or log_date_lis[3:6] == 'Dec':
                    difference_date = 31 - int(log_date_lis[i][:2]) + int(log_date_lis[i+1][:2])
                else:
                    difference_date = 30 - int(log_date_lis[i][:2]) + int(log_date_lis[i+1][:2])
            total_difference_date_lis.append(difference_date)
        print(total_difference_date_lis)
        total_difference_date_lis = sorted(total_difference_date_lis)
        maximum_value = total_difference_date_lis[-1]
    except OSError as e:
        print(f"maximum_value--Error: {e}")
        maximum_value = 100
    finally:
        return  maximum_value


def pickle_read_date_lis(path):
    try:
        with open("/home/c0a21151/pickle/"+path, 'rb') as f:
            date_lis = pickle.load(f)
    except FileNotFoundError:
        print(f"該当するファイルがありません:date_lis")
        date_lis = []
    except OSError as e:
        print(f"pickle_read_date_lis--Error: {e}")
    finally:
        return date_lis

def pickle_read_maximum_value(path):
    try:
        with open("/home/c0a21151/pickle/"+path, 'rb') as f:
            maximum_value = pickle.load(f)
    except FileNotFoundError:
        print(f"該当するファイルがありません:maximum_value")
        maximum_value = 100
    except OSError as e:
        print(f"pickle_read_maximum_value--Error: {e}")
    finally:
        return maximum_value

def pickle_read_no_access_count(path):
    try:
        with open("/home/c0a21151/pickle/"+path, 'rb') as f:
            no_access_count = pickle.load(f)
    except FileNotFoundError:
        print(f"該当するファイルがありません:no_access_count")
        no_access_count = 0
    except OSError as e:
        print(f"pickle_read_no_access_count--Error: {e}")
    finally:
        return no_access_count

def ssh_determine_shutdown(date_lis, old_date_lis, old_no_access_count, frg):
    try:
        date_lis_p = date_lis.copy()
        no_access_count = old_no_access_count
        #date_lis_p.append("Dec 31")
        print(f"テストリスト：{date_lis_p}")

        if old_date_lis == []:
            old_date_lis = date_lis_p
            no_access_count = 0
            frg = 0
        
        # 今日の日付を取得
        today = datetime.now()
        month_name = today.strftime('%B')
        
        if len(today.day) == 1:
            day = f"{month_name[:3]}  {today.day}"
        elif len(today.day) == 2:
            day = f"{month_name[:3]} {today.day}"

        for i in old_date_lis:
            if date_lis_p[0][:3] == i[:3] and date_lis_p[0][-2:] >= i[-2:]:
                a = old_date_lis.pop(0)
                print(f"*******a:{a}")
                continue
            break
            
        if date_lis_p[-1] != day:
            no_access_count += 1
        else:
            """
            for i in old_date_lis:
                for l in range(len(date_lis_p)-1):
                    print(f"i,l, date_lis_p[l]:{i}, {l}, {date_lis_p[l]}")
                    if i == date_lis_p[l]:
                        p = date_lis_p.pop(l)
                        print(f"{p}")
                        break
                print(f"date_lis_p:{date_lis_p}")
            """
            old_date_lis.append(day)
            no_access_count = 0
        #no_access_count = 30
        # if no_access_count <= maximum_value:
        #     print(f"今日の実行終了")
        # else:
        #     frg = 1
        #     print(f"HTTPログをみる...{no_access_count}")
    except OSError as e:
        print(f"determine_shutdown--Error: {e}")
    finally:
        return old_date_lis, no_access_count, frg


def http_determine_shutdown(date_lis, old_date_lis, old_no_access_count, frg):
    try:
        date_lis_p = date_lis.copy()
        no_access_count = old_no_access_count
        #date_lis_p.append("31/Dec/2023")
        print(f"テストリスト：{date_lis_p}")

        if old_date_lis == []:
            old_date_lis = date_lis_p
            no_access_count = 0
            frg = 0

        # 今日の日付を取得
        today = datetime.now()
        month_name = today.strftime('%B')
        
        day = f"{today.day}/{month_name[:3]}/{today.year}"

        for i in old_date_lis:
            if date_lis_p[0][:3] == i[:3] and date_lis_p[0][-2:] >= i[-2:]:
                a = old_date_lis.pop(0)
                print(f"*******a:{a}")
                continue
            break            

        if date_lis_p[-1] != day:
            no_access_count += 1
        else:
            """
            for i in old_date_lis:
                for l in range(len(date_lis_p)-1):
                    print(f"i,l, date_lis_p[l]:{i}, {l}, {date_lis_p[l]}")
                    if i == date_lis_p[l]:
                        p = date_lis_p.pop(l)
                        print(f"{p}")
                        break
                print(f"date_lis_p:{date_lis_p}")
            """
            old_date_lis.append(day)
            no_access_count = 0
        #no_access_count = 30
        # if no_access_count <= maximum_value:
        #     print(f"今日の実行終了")
        # else:
        #     frg = 1
        #     print(f"シャットダウンする...{no_access_count}")
    except OSError as e:
        print(f"determine_shutdown--Error: {e}")
    finally:
        return old_date_lis, 0, frg


def pickle_write(date_lis, no_access_count, path_date_lis, path_no_access_count):
    try:
        print(f"pickle_write:{date_lis}, {no_access_count}")
        #write_lis = [0,1,2]
        with open("/home/c0a21151/pickle/"+path_date_lis, 'wb') as f:
            pickle.dump(date_lis, f)
        #with open('maximum_value.pickle', 'wb') as f:
            #pickle.dump(maximum_value, f)
        with open("/home/c0a21151/pickle/"+path_no_access_count, 'wb') as f:
            pickle.dump(no_access_count, f)
    except OSError as e:
        print(f"pickle_write--Error: {e}")

def ssh_remort_connect(user_name, host_name):
    """
    # SSHの設定
    hostname = 'c0a21151-test0'
    username = 'c0a21151'
    """
    port = 22
    private_key_path = '/home/c0a21151/.ssh/id_ed25519'

    # SSHセッションの確立
    client = paramiko.SSHClient()
    client.load_system_host_keys()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    private_key = paramiko.Ed25519Key(filename=private_key_path)
    client.connect(host_name, port, user_name, pkey=private_key)

    # リモートディレクトリを取得
    remote_directory = '/var/log/'
    command = f"ls -a {remote_directory}"
    stdin, stdout, stderr = client.exec_command(command)

    file_name = stdout.read().decode('utf-8')

    # 結果の表示
    fn = file_name.split("\n")
    #print(fn)
    
    return client, fn

def http_remort_connect(user_name, host_name):
    """
    # HTTPの設定
    hostname = 'c0a21151-test0'
    username = 'c0a21151'
    """
    port = 22
    private_key_path = '/home/c0a21151/.ssh/id_ed25519'

    # SSHセッションの確立
    client = paramiko.SSHClient()
    client.load_system_host_keys()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    private_key = paramiko.Ed25519Key(filename=private_key_path)
    client.connect(host_name, port, user_name, pkey=private_key)

    # リモートディレクトリを取得
    remote_directory = '/var/log/apache2/'
    command = f"ls -a {remote_directory}"
    stdin, stdout, stderr = client.exec_command(command)

    file_name = stdout.read().decode('utf-8')

    # 結果の表示
    fn = file_name.split("\n")
    #print(fn)

    return client, fn


def ssh_log(user_name, host_name):
    client, fn = ssh_remort_connect(user_name, host_name)
    log_dir_path = user_name + "@" + host_name + ":" 

    auth_log_lis = [filename for filename in fn if filename.startswith("auth")]
    print(f"auth_log_lis: {auth_log_lis}")

    #local_dir = "/home/" + user_name + "/auth_log/"
    local_dir = "/home/c0a21151/auth_log/"
    os.makedirs(local_dir, exist_ok=True)

    remote_directory = "/var/log/"
    file_lis_gz = []
    # gzipファイルをリモートからローカルに転送
    for i in auth_log_lis:
        #print("00")
        if i[-3:] == ".gz":
            #print(f"i, i[-3:]: {i, i[-3:]}")
            extract_file_path = local_dir + i[:-3]
            #print(f"extract_file_path:{extract_file_path}")
            extract_gzip(remote_directory + i, extract_file_path, client)
            file_lis_gz.append(i[:-3])
        else:
            #print("22")
            with client.open_sftp() as sftp:
                sftp.get(remote_directory + i, local_dir + i)
            file_lis_gz.append(i)
    #print(f"123: {file_lis_gz}")

    date_lis = []
    date_lis = ssh_sort_date_lis(file_lis_gz, local_dir, date_lis)
    date_count = len(date_lis)
    directory_rm(local_dir)
    
    print(date_lis)

    if date_lis == []:
        print("リストが空です")
        shutdown_frg = 1
        return shutdown_frg, 0
    if len(date_lis) == 1:
        print("アクセス日が1日のみです")
        shutdown_frg = 0
        return shutdown_frg, 0

    #maximum_value = ssh_maximum_value_access_frequency(date_lis)

    #Average_access_frequency = ssh_Average_access_frequency_calculation(date_lis, date_count)
    #Average_access_frequency = round(Average_access_frequency)

    # print(f"日数：{date_count}")
    # print(f"アクセス日：{date_lis}")
    # print(f"アクセス間隔最大値：{maximum_value}")
    #print(f"平均アクセス頻度：{Average_access_frequency}")
    
    print(f"アクセス最終日：{date_lis[-1]}")

    old_date_lis =  pickle_read_date_lis(host_name + '_ssh_date_lis.pickle')
    #old_maximum_value =  pickle_read_maximum_value(host_name + '_ssh_maximum_value.pickle')
    old_no_access_count = pickle_read_no_access_count(host_name + '_ssh_no_access_count.pickle')
    print(f"old_date_lis, old_no_access_count:{old_date_lis}, {old_no_access_count}")

    http_frg = 0
    new_date_lis, new_no_access_count, http_frg = ssh_determine_shutdown(date_lis, old_date_lis, old_no_access_count, http_frg)
    print(f"new_date_lis:{new_date_lis}")
    print(f"new_no_access_count:{new_no_access_count}")
    print(f"http_frg:{http_frg}")

    maximum_value = ssh_maximum_value_access_frequency(new_date_lis)
    print(f"アクセス間隔最大値：{maximum_value}")

    pickle_write(new_date_lis, new_no_access_count, host_name + '_ssh_date_lis.pickle', host_name + '_ssh_no_access_count.pickle')
    
    http_frg = 0

    # SSHセッションのクローズ
    client.close()

    return http_frg, maximum_value

    
def http_log(user_name, host_name):
    client, fn = http_remort_connect(user_name, host_name)
    log_dir_path = "user_name" + "@" + host_name + ":"
    
    access_log_lis = [filename for filename in fn if filename.startswith("access")]
    print(f"access_log_lis: {access_log_lis}")

    local_dir = "/home/c0a21151/access_log/"
    os.makedirs(local_dir, exist_ok=True)

    remote_directory = "/var/log/apache2/"
    file_lis_gz = []
    # gzipファイルをリモートからローカルに転送
    for i in access_log_lis:
        #print("00")
        if i[-3:] == ".gz":
            print(f"i, i[-3:]: {i, i[-3:]}")
            extract_file_path = local_dir + i[:-3]
            #print(f"extract_file_path:{extract_file_path}")
            extract_gzip(remote_directory + i, extract_file_path, client)
            file_lis_gz.append(i[:-3])
        else:
            #print("22")
            with client.open_sftp() as sftp:
                sftp.get(remote_directory + i, local_dir + i)
            file_lis_gz.append(i)
    #print(f"123: {file_lis_gz}")

    date_lis = []
    date_lis = http_sort_date_lis(file_lis_gz, local_dir, date_lis)
    date_count = len(date_lis)
    directory_rm(local_dir)

    print(f"546,,date_lis: {date_lis}")
    if date_lis == []:
        print("リストが空です")
        shutdown_frg = 1
        return shutdown_frg, 0
    if len(date_lis) == 1:
        print("アクセス日が1日のみです")
        shutdown_frg = 0
        return shutdown_frg, 0

    #maximum_value = http_maximum_value_access_frequency(date_lis)

    #Average_access_frequency = Average_access_frequency_calculation(date_lis, date_count)
    #Average_access_frequency = round(Average_access_frequency)

    # print(f"日数：{date_count}")
    # print(f"アクセス日：{date_lis}")
    # print(f"アクセス間隔最大値：{maximum_value}")
    #print(f"平均アクセス頻度：{Average_access_frequency}")
    
    print(f"アクセス最終日：{date_lis[-1]}")

    old_date_lis = pickle_read_date_lis(host_name + '_http_date_lis.pickle')
    #old_maximum_value =  pickle_read_maximum_value(host_name + '_http_maximum_value.pickle')
    old_no_access_count = pickle_read_no_access_count(host_name + '_http_no_access_count.pickle')
    print(f"old_date_lis, old_no_access_count:{old_date_lis}, {old_no_access_count}")

    shutdown_frg = 0
    new_date_lis, new_no_access_count, shutdown_frg = http_determine_shutdown(date_lis, old_date_lis, old_no_access_count, shutdown_frg)
    print(f"new_date_lis:{new_date_lis}")
    print(f"new_no_access_count:{new_no_access_count}")
    print(f"shutdown_frg:{shutdown_frg}")

    maximum_value = http_maximum_value_access_frequency(new_date_lis)
    print(f"アクセス間隔最大値：{maximum_value}")

    pickle_write(new_date_lis, new_no_access_count, host_name + '_http_date_lis.pickle', host_name + '_http_no_access_count.pickle')
    
    shutdown_frg = 0

    # SSHセッションのクローズ
    client.close()

    return shutdown_frg, maximum_value


def md(maximum_value_lis):
    most = 0
    max = 0
    n_lis = list(set(maximum_value_lis))

    for n in n_lis:
        #取り出したデータと等しいデータの数を数える
        num = maximum_value_lis.count(n)
        #これまでのmaxよりも多い出現回数ならmost, maxを更新
        if num > max:
            max = num
            #最頻値を更新
            most = n
        elif num == max:
            if n > most:
                most = n
    return most


def vm_reference(vm_excel):
    wb = openpyxl.load_workbook(vm_excel)
    ws = wb.worksheets[0]

    user_lis = []
    host_lis = []
    ok_lis = []
    user_lis = [cell.value for cell in ws['B']]
    host_lis = [cell.value for cell in ws['C']]
    vm_pass_lis = [cell.value for cell in ws['D']]
    ok_lis = [cell.value for cell in ws['H']]
    kakunin = [cell.value for cell in ws['J']]
    
    return user_lis, host_lis, vm_pass_lis, ok_lis, kakunin


if __name__ == "__main__":
    vm_excel = 'vm_log_reference.xlsx'
    user_lis, host_lis, vm_pass_lis, ok_lis, kakunin = vm_reference(vm_excel)
    print(user_lis, host_lis, vm_pass_lis, ok_lis, kakunin)

    n = 0
    maximum_value_vm_lis = []
    maximum_value_lis = []

    for i in range(1, len(user_lis)):
        try:
            n += 1
            user_name = user_lis[i]
            host_name = host_lis[i]
            vm_pass = vm_pass_lis[i]
            ok = ok_lis[i]
            print("----- next user -----")
            print(user_name, host_name, ok)

            if ok == None:
                continue

            command = f'ssh -i /home/c0a21151/.ssh/id_ed25519 {user_name}@{host_name} "pwd"'
            result = subprocess.run(command, shell=True, capture_output=True, text=True)
            print(f"result.stderr:{result.stderr}")
            if re.search(r"ssh", result.stderr):
                print("接続できない")
                continue
        
            """
            SSHログを確認する
            """
            frg_http, maximum_value = ssh_log(user_name, host_name)

            maximum_value_vm_lis.append([host_name, maximum_value])
            maximum_value_lis.append(maximum_value)

            if frg_http == 0:
                print(f"起動継続")
                #raise SystemExit
                continue

            print("--------------------")
    
            """
            HTTPログを確認する
            """
            frg_shutdown, maximum_value = http_log(user_name, host_name)
            
            if maximum_value != 0:
                maximum_value_vm_lis.append([host_name, maximum_value])
                maximum_value_lis.append(maximum_value)
            """
            if frg_shutdown == 0:
                print(f"起動継続")
                #raise SystemExit
                maximum_value_lis.append(maximum_value)
                continue
            """
            print("--------------------")

            
        except OSError as e:
            print(f"main--Error: {e}")
            print(f"user_name, host_name: {user_name}, {host_name}")
            continue

    print("--------------------")

    print(f"最大アクセス間隔リスト：{maximum_value_lis}")
    most_maximum_value = md(maximum_value_lis)
    print(f"最頻値：{most_maximum_value}")

    shutdown_vm = []
    for i in maximum_value_vm_lis:
        maximum_value = i[1]
        if maximum_value > most_maximum_value:
            shutdown_vm.append(i[0])
    print(f"shutdown_vm：{shutdown_vm}")    

    for i in range(1, len(user_lis)):
        user_name = user_lis[i]
        host_name = host_lis[i]
        vm_pass = vm_pass_lis[i]
        for j in shutdown_vm:
            if j == host_name:
                """
                シャットダウンする
                """
                print(f"シャットダウンする,,{host_name}, {vm_pass}")
                """
                # user_name = "c0a21037"
                # host_name = "192.168.100.226"
                # vm_pass = "vm2"

                # 実行したいコマンド
                command = f'ssh -i /home/c0a21151/.ssh/id_ed25519 {user_name}@{host_name} "echo {vm_pass} | sudo -S shutdown -h now"'
                #print(command)

                # コマンドを実行して標準出力を取得
                result = subprocess.run(command, shell=True, capture_output=True, text=True)

                # 結果を表示
                if result.returncode == 0:
                    print("シャットダウンしました")
                    
                else:
                    print(f"コマンドはエラーで終了しました。エラーコード: {result.returncode}")
                    print(f"標準エラー: {result.stderr}")
                """

    print(f"回転数：{n}")


